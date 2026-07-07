import sys
import os
import glob
import math
from datetime import datetime
from PyQt6.QtGui import QPixmap

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFormLayout, QComboBox, QSpinBox,
    QListWidget, QMessageBox, QFileDialog, QRadioButton, QButtonGroup,
    QTextEdit, QGroupBox
)
from PyQt6.QtCore import Qt


ENEMIGOS_COMUNES = {}

def load_enemigos_from_excel(path):
    """
    Carga enemigos desde un archivo Excel (.xlsx)
    devuelve dict {nombre: {"xp_unidad": xp}}.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        raise ImportError("Instala openpyxl: python -m pip install openpyxl")

    def texto(value):
        return "" if value is None else str(value).strip()

    def parse_xp(value):
        if value is None:
            return None
        text = str(value).strip()
        if text == "":
            return None
        try:
            return int(float(text))
        except Exception:
            return None

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    encabezados = [texto(c).lower() for c in rows[0]]
    tiene_header = any(
        ("nombre" in c or "enemigo" in c or "monster" in c or "creature" in c or "name" in c)
        for c in encabezados
    ) and any(
        ("xp" in c or "experiencia" in c or "exp" in c)
        for c in encabezados
    )

    name_idx = 0
    xp_idx = 1

    if tiene_header:
        for i, celda in enumerate(encabezados):
            if any(k in celda for k in ("nombre", "enemigo", "monster", "creature", "name")):
                name_idx = i
            if any(k in celda for k in ("xp", "experiencia", "exp")):
                xp_idx = i

    enemigos = {}
    start = 1 if tiene_header else 0

    for fila in rows[start:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in fila):
            continue

        nombre = fila[name_idx] if name_idx < len(fila) else None
        if not nombre:
            continue
        nombre = str(nombre).strip()

        xp_val = fila[xp_idx] if xp_idx < len(fila) else None
        xp = parse_xp(xp_val)
        if xp is None:
            xp = 0

        enemigos[nombre] = {"xp_unidad": xp}

    return enemigos


ORO_BASE = {
    "🐺 Lobo": 50,
    "🦅 Grifo": 100,
    "🦁 Quimera": 200,
    "🐉 Hidra": 500,
    "🐲 Dragón": 1000
}

TIPOS_MISION = {
    "Misión Estándar / One Shot": {"xp_mod": 1.0, "oro_mod": 1.0, "xp_fija": None},
    "Misión Rolística": {"xp_mod": 0.0, "oro_mod": 1.0, "xp_fija": 50},
    "Misión Rápida / Contrarreloj": {"xp_mod": 0.8, "oro_mod": 1.2, "xp_fija": None},
    "Minicampaña": {"xp_mod": 1.3, "oro_mod": 3.0, "xp_fija": None},
    "Dungeon Crawler": {"xp_mod": 1.0, "oro_mod": 4.0, "xp_fija": None}
}

LEVEL_XP_THRESHOLDS = {
    1: {"Rolística": 25, "Rápida": 50, "Moderado": 75, "Desafiante": 100, "Extremo": 125},
    2: {"Rolística": 50, "Rápida": 100, "Moderado": 150, "Desafiante": 200, "Extremo": 250},
    3: {"Rolística": 100, "Rápida": 200, "Moderado": 300, "Desafiante": 400, "Extremo": 500},
    4: {"Rolística": 175, "Rápida": 350, "Moderado": 525, "Desafiante": 700, "Extremo": 875},
    5: {"Rolística": 275, "Rápida": 550, "Moderado": 825, "Desafiante": 1100, "Extremo": 1375},
    6: {"Rolística": 400, "Rápida": 800, "Moderado": 1200, "Desafiante": 1600, "Extremo": 2000},
    7: {"Rolística": 550, "Rápida": 1100, "Moderado": 1650, "Desafiante": 2200, "Extremo": 2750},
    8: {"Rolística": 750, "Rápida": 1500, "Moderado": 2250, "Desafiante": 3000, "Extremo": 3750},
    9: {"Rolística": 850, "Rápida": 1700, "Moderado": 2550, "Desafiante": 3400, "Extremo": 4250},
    10: {"Rolística": 1100, "Rápida": 2200, "Moderado": 3300, "Desafiante": 4400, "Extremo": 5500},
    11: {"Rolística": 1400, "Rápida": 2800, "Moderado": 4200, "Desafiante": 5600, "Extremo": 7000},
    12: {"Rolística": 1750, "Rápida": 3500, "Moderado": 5250, "Desafiante": 7000, "Extremo": 8750},
    13: {"Rolística": 2200, "Rápida": 4400, "Moderado": 6600, "Desafiante": 8800, "Extremo": 11000},
    14: {"Rolística": 2700, "Rápida": 5400, "Moderado": 8100, "Desafiante": 10800, "Extremo": 13500},
    15: {"Rolística": 3250, "Rápida": 6500, "Moderado": 9750, "Desafiante": 13000, "Extremo": 16250},
    16: {"Rolística": 3850, "Rápida": 7700, "Moderado": 11550, "Desafiante": 15400, "Extremo": 19250},
    17: {"Rolística": 4500, "Rápida": 9000, "Moderado": 13500, "Desafiante": 18000, "Extremo": 22500},
    18: {"Rolística": 5200, "Rápida": 10400, "Moderado": 15600, "Desafiante": 20800, "Extremo": 26000},
    19: {"Rolística": 5950, "Rápida": 11900, "Moderado": 17850, "Desafiante": 23800, "Extremo": 29750},
    20: {"Rolística": 6750, "Rápida": 13500, "Moderado": 20250, "Desafiante": 27000, "Extremo": 33750}
}


class DNDRewardApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("D&D Resumen de Recompensas")
        self.setGeometry(100, 100, 920, 660)

        self.enemigos_lista = []
        self.jugadores_lista = []
        self.dragon_pixmap = self.load_dragon_pixmap()

        self.init_ui()
        self.cargar_excel_default()

    def load_dragon_pixmap(self):
        path = r"C:\Users\migue\OneDrive\Escritorio\pruebas codigos\dragon.png"
        if os.path.exists(path):
            return QPixmap(path)
        return None

    def init_ui(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #1E1515;
            }
            QTabWidget::pane {
                border: 2px solid #8B6914;
                background-color: #2A1112;
                border-radius: 5px;
            }
            QTabBar::tab {
                background: #3E2723;
                color: #C5A059;
                border: 2px solid #8B6914;
                padding: 8px 25px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-family: "Georgia", serif;
                font-size: 14px;
            }
            QTabBar::tab:selected {
                background: #8B6914;
                color: #1E1515;
                font-weight: bold;
            }
            QLabel {
                color: #F4E8C1;
                font-family: "Georgia", serif;
                font-size: 14px;
                font-weight: bold;
            }
            QLineEdit, QComboBox, QSpinBox, QListWidget, QTextEdit {
                background-color: #F4E8C1;
                color: #2B2B2B;
                border: 2px solid #8B6914;
                border-radius: 4px;
                padding: 5px;
                font-family: "Georgia", serif;
                font-size: 13px;
            }
            QComboBox QAbstractItemView {
                background-color: #F4E8C1;
                color: #2B2B2B;
                selection-background-color: #8B6914;
                selection-color: #1E1515;
            }
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QTextEdit:focus {
                border: 2px solid #FFD700;
            }
            QPushButton {
                background-color: #721C1C;
                color: #F4E8C1;
                border: 2px outset #C5A059;
                border-radius: 8px;
                padding: 10px 20px;
                font-family: "Georgia", serif;
                font-weight: bold;
                font-size: 15px;
            }
            QPushButton:hover {
                background-color: #8B2525;
                border: 2px solid #FFD700;
            }
            QPushButton:pressed {
                background-color: #4A1212;
                border: 2px inset #8B6914;
            }
            QGroupBox {
                border: 2px solid #8B6914;
                border-radius: 6px;
                margin-top: 10px;
                color: #F4E8C1;
                font-family: "Georgia", serif;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px 0 3px;
            }
        """)

        tabs = QTabWidget()

        tabs.addTab(self.build_mission_tab(), "Detalles de Misión")
        tabs.addTab(self.build_enemies_tab(), "Enemigos Derrotados")
        tabs.addTab(self.build_players_tab(), "Jugadores")

        btn_report = QPushButton("⚔️ GENERAR REPORTE DE DISCORD")
        btn_report.clicked.connect(self.generar_reporte)

        footer = QWidget()
        footer_layout = QHBoxLayout(footer)
        footer_layout.addStretch()
        footer_layout.addWidget(btn_report)
        footer_layout.addStretch()

        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)
        main_layout.addWidget(tabs)
        main_layout.addWidget(footer)

        self.setCentralWidget(main_widget)

    def build_mission_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        header = QWidget()
        header_layout = QHBoxLayout(header)

        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        self.session_name = QLineEdit()
        self.session_name.setText("Limpieza del cementerio")
        self.session_date = QLineEdit()
        self.session_date.setText(datetime.now().strftime("%d/%m/%Y"))
        self.session_type = QComboBox()
        self.session_type.addItems(list(TIPOS_MISION.keys()))
        self.session_type.setCurrentIndex(0)

        form_layout.addRow(QLabel("Nombre de la Sesión:"), self.session_name)
        form_layout.addRow(QLabel("Fecha (DD/MM/YYYY):"), self.session_date)
        form_layout.addRow(QLabel("Tipo de Misión:"), self.session_type)

        left_widget = QWidget()
        left_widget.setLayout(form_layout)
        header_layout.addWidget(left_widget)

        if self.dragon_pixmap is not None:
            dragon_label = QLabel()
            dragon_label.setPixmap(self.dragon_pixmap)
            dragon_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
            header_layout.addWidget(dragon_label)

        layout.addWidget(header)
        layout.addStretch()

        return tab

    def build_enemies_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        form_frame = QWidget()
        form_layout = QHBoxLayout(form_frame)

        left_form = QWidget()
        left_layout = QFormLayout(left_form)
        left_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.combo_nombre_enemigo = QComboBox()
        self.combo_nombre_enemigo.setEditable(True)
        self.combo_nombre_enemigo.currentTextChanged.connect(self.mostrar_xp_enemigo)
        self.combo_nombre_enemigo.lineEdit().setPlaceholderText("Escribe o selecciona enemigo")

        self.label_xp_enemigo = QLabel("0")
        self.spin_cant_enemigo = QSpinBox()
        self.spin_cant_enemigo.setRange(1, 100)
        self.spin_cant_enemigo.setValue(1)

        left_layout.addRow(QLabel("Enemigo:"), self.combo_nombre_enemigo)
        left_layout.addRow(QLabel("XP por unidad:"), self.label_xp_enemigo)
        left_layout.addRow(QLabel("Cantidad:"), self.spin_cant_enemigo)

        form_layout.addWidget(left_form)

        botones_frame = QWidget()
        botones_layout = QVBoxLayout(botones_frame)
        botones_layout.setSpacing(10)
        botones_layout.addWidget(QPushButton("Añadir Enemigo", clicked=self.agregar_enemigo))
        botones_layout.addWidget(QPushButton("Importar Excel", clicked=self.importar_excel_enemigos))
        botones_layout.addStretch()

        form_layout.addWidget(botones_frame)
        layout.addWidget(form_frame)

        self.label_status_enemigos = QLabel("")
        layout.addWidget(self.label_status_enemigos)

        self.listbox_enemigos = QListWidget()
        layout.addWidget(self.listbox_enemigos)

        remove_btn = QPushButton("Eliminar Seleccionado")
        remove_btn.clicked.connect(self.eliminar_enemigo)
        layout.addWidget(remove_btn, alignment=Qt.AlignmentFlag.AlignRight)

        return tab

    def build_players_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        form_group = QGroupBox("Datos del Jugador")
        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.player_name = QLineEdit()
        self.player_level = QComboBox()
        self.player_level.addItems([str(i) for i in range(1, 21)])
        self.player_level.setCurrentIndex(0)

        self.player_rank = QComboBox()
        self.player_rank.addItems(list(ORO_BASE.keys()))
        self.player_rank.setCurrentIndex(0)

        self.player_difficulty = QComboBox()
        self.player_difficulty.addItems(["Rolística", "Rápida", "Moderado", "Desafiante", "Extremo"])
        self.player_difficulty.setCurrentIndex(2)

        self.boost_no = QRadioButton("No")
        self.boost_si = QRadioButton("Sí")
        self.boost_no.setChecked(True)
        boost_group = QButtonGroup()
        boost_group.addButton(self.boost_no)
        boost_group.addButton(self.boost_si)

        boost_widget = QWidget()
        boost_layout = QHBoxLayout(boost_widget)
        boost_layout.addWidget(self.boost_no)
        boost_layout.addWidget(self.boost_si)

        self.boost_type = QComboBox()
        self.boost_type.addItems(["Dinero", "Experiencia", "Ambas"])
        self.boost_type.setCurrentIndex(2)

        self.spin_porcentaje_boost = QSpinBox()
        self.spin_porcentaje_boost.setRange(1, 100)
        self.spin_porcentaje_boost.setValue(10)

        self.entry_motivo_boost = QLineEdit()

        form_layout.addRow(QLabel("Mención completa:"), self.player_name)
        form_layout.addRow(QLabel("Nivel del jugador:"), self.player_level)
        form_layout.addRow(QLabel("Dificultad de la misión:"), self.player_difficulty)
        form_layout.addRow(QLabel("Rango de Aventurero:"), self.player_rank)
        form_layout.addRow(QLabel("¿Tiene Boost?"), boost_widget)
        form_layout.addRow(QLabel("Tipo de Boost:"), self.boost_type)
        form_layout.addRow(QLabel("% Boost:"), self.spin_porcentaje_boost)
        form_layout.addRow(QLabel("Motivo:"), self.entry_motivo_boost)
        form_group.setLayout(form_layout)

        layout.addWidget(form_group)

        btn_add = QPushButton("Añadir Jugador")
        btn_add.clicked.connect(self.agregar_jugador)
        layout.addWidget(btn_add, alignment=Qt.AlignmentFlag.AlignLeft)

        self.listbox_jugadores = QListWidget()
        layout.addWidget(self.listbox_jugadores)

        remove_btn = QPushButton("Eliminar Seleccionado")
        remove_btn.clicked.connect(self.eliminar_jugador)
        layout.addWidget(remove_btn, alignment=Qt.AlignmentFlag.AlignRight)

        return tab

    def actualizar_combobox_enemigos(self):
        self.combo_nombre_enemigo.clear()
        self.combo_nombre_enemigo.addItems(sorted(ENEMIGOS_COMUNES.keys()))

    def mostrar_xp_enemigo(self, text):
        xp = 0
        if text in ENEMIGOS_COMUNES:
            xp = ENEMIGOS_COMUNES[text].get("xp_unidad", 0)
        self.label_xp_enemigo.setText(str(xp))

    def agregar_enemigo(self):
        nombre = self.combo_nombre_enemigo.currentText().strip()
        cantidad = self.spin_cant_enemigo.value()

        if not nombre:
            QMessageBox.warning(self, "Error", "Falta el nombre del enemigo.")
            return

        xp_unidad = 0
        if nombre in ENEMIGOS_COMUNES:
            xp_unidad = ENEMIGOS_COMUNES[nombre].get("xp_unidad", 0)

        enemigo_data = {
            "nombre": nombre,
            "cantidad": cantidad,
            "xp_unidad": xp_unidad,
            "xp_total": xp_unidad * cantidad
        }
        self.enemigos_lista.append(enemigo_data)
        self.listbox_enemigos.addItem(f"x{cantidad} {nombre} | {xp_unidad} xp c/u")
        self.combo_nombre_enemigo.setCurrentText("")
        self.label_xp_enemigo.setText("0")
        self.spin_cant_enemigo.setValue(1)

    def eliminar_enemigo(self):
        row = self.listbox_enemigos.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Aviso", "No hay enemigo seleccionado.")
            return
        self.enemigos_lista.pop(row)
        self.listbox_enemigos.takeItem(row)

    def agregar_jugador(self):
        nombre_completo = self.player_name.text().strip()
        if not nombre_completo:
            QMessageBox.warning(self, "Error", "Escribe el nombre del jugador.")
            return

        nivel = int(self.player_level.currentText())
        dificultad = self.player_difficulty.currentText()
        rango = self.player_rank.currentText()
        tiene_boost = self.boost_si.isChecked()
        tipo_boost = self.boost_type.currentText()
        motivo_boost = self.entry_motivo_boost.text().strip()

        nombre_corto = nombre_completo.split("|")[0].strip() if "|" in nombre_completo else nombre_completo.split()[0]
        porcentaje_boost = self.spin_porcentaje_boost.value()

        jugador_data = {
            "nombre_completo": nombre_completo,
            "nombre_corto": nombre_corto,
            "nivel": nivel,
            "dificultad": dificultad,
            "rango": rango,
            "boost": "Sí" if tiene_boost else "No",
            "tipo_boost": tipo_boost if tiene_boost else None,
            "porcentaje": porcentaje_boost if tiene_boost else 0,
            "motivo_boost": motivo_boost if tiene_boost else ""
        }

        self.jugadores_lista.append(jugador_data)
        self.listbox_jugadores.addItem(f"{nombre_corto} (Nv {nivel}, {dificultad}) añadido.")
        self.player_name.clear()
        self.entry_motivo_boost.clear()

    def eliminar_jugador(self):
        row = self.listbox_jugadores.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Aviso", "No hay jugador seleccionado.")
            return
        self.jugadores_lista.pop(row)
        self.listbox_jugadores.takeItem(row)

    def importar_excel_enemigos(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Excel de Enemigos",
            "",
            "Excel files (*.xlsx *.xlsm *.xltx *.xltm);;All files (*)"
        )
        if not path:
            return

        try:
            enemigos = load_enemigos_from_excel(path)
        except ImportError as ie:
            QMessageBox.critical(self, "Dependencia faltante", str(ie))
            return
        except Exception as e:
            self.label_status_enemigos.setText(f"Error leyendo Excel: {e}")
            return

        if enemigos:
            global ENEMIGOS_COMUNES
            ENEMIGOS_COMUNES = enemigos
            self.actualizar_combobox_enemigos()
            self.label_status_enemigos.setText(f"Cargados {len(enemigos)} enemigos (Excel)")
        else:
            self.label_status_enemigos.setText("Excel leído: no se encontraron enemigos válidos")

    def cargar_excel_default(self):
        carpeta = r"C:\Users\migue\OneDrive\Escritorio\pruebas codigos"
        archivos = glob.glob(os.path.join(carpeta, "*.xlsx"))

        if not archivos:
            self.label_status_enemigos.setText("No se encontró Excel de enemigos")
            return

        path = archivos[0]
        try:
            enemigos = load_enemigos_from_excel(path)
        except ImportError:
            self.label_status_enemigos.setText("Falta openpyxl para cargar Excel")
            return
        except Exception as e:
            self.label_status_enemigos.setText(f"Error leyendo Excel: {e}")
            return

        if enemigos:
            global ENEMIGOS_COMUNES
            ENEMIGOS_COMUNES = enemigos
            self.actualizar_combobox_enemigos()
            self.label_status_enemigos.setText(f"Cargados {len(enemigos)} enemigos (Excel)")
        else:
            self.label_status_enemigos.setText("Excel leído: no se encontraron enemigos válidos")

    def generar_reporte(self):
        if not self.jugadores_lista:
            QMessageBox.warning(self, "Aviso", "No hay jugadores añadidos.")
            return

        tipo_mision = self.session_type.currentText()
        reglas_mision = TIPOS_MISION[tipo_mision]

        xp_total_monstruos = sum(e["xp_total"] for e in self.enemigos_lista)

        reporte = f"{self.session_name.text()}\n"
        reporte += f"{self.session_date.text()}\n\n\n"
        reporte += "⚔️ Jugadores ⚔️\n"
        for j in self.jugadores_lista:
            reporte += f"{j['nombre_completo']}\n"

        reporte += "\n🧠 Experiencia\n"
        reporte += f"Total XP de monstruos: {int(xp_total_monstruos)} xp\n\n"
        reporte += "🥊 Enemigos Superados\n"
        for e in self.enemigos_lista:
            reporte += f"x{e['cantidad']} {e['nombre']} | {e['xp_unidad']} xp c/u\n"

        reporte += "\nRecompensas\n"

        grupos_oro = {}
        grupos_xp = {}

        cant_jugadores = len(self.jugadores_lista)
        xp_por_nivel = []
        if cant_jugadores > 0:
            if reglas_mision["xp_fija"] is not None:
                xp_base = reglas_mision["xp_fija"]
                xp_por_nivel = [xp_base] * cant_jugadores
            else:
                for j in self.jugadores_lista:
                    nivel = j.get("nivel", 1)
                    dificultad = j.get("dificultad", "Moderado")
                    xp_por_nivel.append(LEVEL_XP_THRESHOLDS.get(nivel, LEVEL_XP_THRESHOLDS[1]).get(dificultad, 0))
                total_xp_nivel = sum(xp_por_nivel)
                if total_xp_nivel > 0:
                    escala = (xp_total_monstruos * reglas_mision["xp_mod"]) / total_xp_nivel
                    xp_por_nivel = [math.floor(x * escala) for x in xp_por_nivel]
        else:
            xp_por_nivel = []

        for idx, j in enumerate(self.jugadores_lista):
            base_oro = ORO_BASE.get(j["rango"], 0)
            oro_final = base_oro * reglas_mision["oro_mod"]
            xp_final = xp_por_nivel[idx] if idx < len(xp_por_nivel) else 0
            motivo_formateado = ""

            if j["boost"] == "Sí":
                bono = j["porcentaje"] / 100.0
                if j["tipo_boost"] in ["Dinero", "Ambas"]:
                    oro_final += base_oro * bono
                if j["tipo_boost"] in ["Experiencia", "Ambas"]:
                    xp_final += math.floor((xp_por_nivel[idx] if idx < len(xp_por_nivel) else 0) * bono)

                if j["motivo_boost"]:
                    motivo_formateado = f" ({j['motivo_boost']})"
                else:
                    motivo_formateado = f" ({j['porcentaje']}% extra)"

            xp_final = math.floor(xp_final)

            clave_oro = (int(oro_final), motivo_formateado if j["tipo_boost"] in ["Dinero", "Ambas"] else "")
            grupos_oro.setdefault(clave_oro, []).append(j["nombre_corto"])

            clave_xp = (xp_final, motivo_formateado if j["tipo_boost"] in ["Experiencia", "Ambas"] else "")
            grupos_xp.setdefault(clave_xp, []).append(j["nombre_corto"])

        for (oro_val, motivo), nombres in grupos_oro.items():
            nombres_str = " - ".join(nombres)
            reporte += f"{oro_val} 🪙 {nombres_str}{motivo}\n"

        for (xp_val, motivo), nombres in grupos_xp.items():
            nombres_str = " - ".join(nombres)
            reporte += f"{xp_val} xp {nombres_str}{motivo}\n"

        ventana_reporte = QWidget()
        ventana_reporte.setWindowTitle("Copia tu Reporte")
        layout = QVBoxLayout(ventana_reporte)
        text_area = QTextEdit()
        text_area.setReadOnly(True)
        text_area.setPlainText(reporte)
        layout.addWidget(text_area)
        ventana_reporte.resize(640, 520)
        ventana_reporte.show()

        self.reporte_window = ventana_reporte


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DNDRewardApp()
    window.show()
    sys.exit(app.exec())